import openpyxl
import openpyxl.styles
import tkinter.ttk as ttk
import tkinter as tk
import time
from tkinter import filedialog, messagebox, simpledialog
from tkinter.ttk import Combobox, Style
from tkcalendar import DateEntry
from datetime import datetime

import sys
from sys import exit

import os
from pathlib import Path

def center_window(window):
    window.update_idletasks()
    width = window.winfo_width()
    height = window.winfo_height()
    x = (window.winfo_screenwidth() // 2) - (width // 2)
    y = (window.winfo_screenheight() // 2) - (height // 2)
    window.geometry('{}x{}+{}+{}'.format(width, height, x, y))

root = tk.Tk()
root.geometry('500x400')
root.title('費用入力ツール')

if getattr(sys, 'frozen', False):
    # If the application is running in a bundle, use the bundle resources path
    root.iconbitmap(default=sys._MEIPASS + '/icon.ico')
else:
    # If the application is not running in a bundle, use the local resources path
    root.iconbitmap(default='icon.ico')

center_window(root)

input_frame = None  # 声明全局变量 input_frame
worksheet_name_combobox = None
new_button = None
ok_button = None
reload_button = None


my_style = Style()
my_style.configure('my.TFrame', background='#ececec')
my_style.configure('my.TCombobox', width=20, padding=(5, 5, 5, 5), foreground='black', background='white', font=('Arial', 14))
my_style.configure('my.TEntry', width=20, padding=(5, 5, 5, 5), foreground='black', background='white', font=('Arial', 14))
my_style.configure('my.TButton', padding=(5, 5, 5, 5), foreground='black', background='white')

def get_user_record_folder():
    documents_folder = str(Path.home() / "Documents")
    user_record_folder = os.path.join(documents_folder, "費用入力保存資料")
    
    if not os.path.exists(user_record_folder):
        os.makedirs(user_record_folder)

    return user_record_folder

def save_file_path(file_path):
    save_path = os.path.join(get_user_record_folder(), 'file_path.txt')
    with open(save_path, 'w') as f:
        f.write(file_path)

def load_file_path():
    load_path = os.path.join(get_user_record_folder(), 'file_path.txt')
    if os.path.exists(load_path):
        with open(load_path, 'r') as f:
            file_path = f.read().strip()
            if os.path.exists(file_path):
                return file_path
    return None

def save_last_worksheet_name(worksheet_name):
    save_path = os.path.join(get_user_record_folder(), 'last_worksheet_name.txt')
    with open(save_path, 'w') as f:
        f.write(worksheet_name)

def load_last_worksheet_name():
    load_path = os.path.join(get_user_record_folder(), 'last_worksheet_name.txt')
    if os.path.exists(load_path):
        with open(load_path, 'r') as f:
            worksheet_name = f.read().strip()
            return worksheet_name
    return None

def is_file_open(file_path):
    while True:
        try:
            with open(file_path, 'r+b') as f:
                return False
        except PermissionError:
            messagebox.showerror("エラーメッセージ", "Excelを閉じてください!")
            time.sleep(1)


# 获取文件路径
file_path = load_file_path()
if not file_path:
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    if not file_path:
        root.destroy()
        exit()
    save_file_path(file_path)
    workbook = openpyxl.load_workbook(file_path)
else:
    workbook = openpyxl.load_workbook(file_path)
    worksheet_names = workbook.sheetnames

# 在此处添加检查文件是否已打开的代码
if is_file_open(file_path):
    messagebox.showerror("エラーメッセージ", "Excelを閉じてください!")
    root.destroy()
    exit()


# 获取上次选择的工作表名称
last_worksheet_name = load_last_worksheet_name()

# 修改原有的下拉选择框创建代码，使用新的 default_worksheet_name 变量作为默认值
worksheet_names = workbook.sheetnames
if last_worksheet_name and last_worksheet_name in worksheet_names:
    default_worksheet_name = last_worksheet_name
else:
    default_worksheet_name = 'シートを選択してください。'

worksheet_name_var = tk.StringVar(value=default_worksheet_name)
worksheet_name_combobox = Combobox(root, style='my.TCombobox', textvariable=worksheet_name_var, values=worksheet_names, width=30)
worksheet_name_combobox.pack(pady=10)



# 创建新工作表
def add_worksheet():
    global worksheet_names
    worksheet_name = simpledialog.askstring("新しいシート", "新しいシート名を入力してください。")
    if worksheet_name:
        master_worksheet = workbook['マスター']
        worksheet = workbook.copy_worksheet(master_worksheet)
        worksheet.title = worksheet_name
        
        # 添加 try-except 块来捕获 PermissionError 异常
        try:
            workbook.save(file_path)
        except PermissionError:
            messagebox.showerror("エラーメッセージ", "Excelを閉じってください!")
            return
        
        worksheet_names = workbook.sheetnames
        worksheet_name_combobox.configure(values=worksheet_names)
        worksheet_name_var.set(worksheet_name)

        # 更新工作表名称列表
        worksheet_names = workbook.sheetnames

        # 将新工作表设置为选中状态
        worksheet_name_var.set(worksheet_name)

        # 设置当前工作表
        worksheet = workbook[worksheet_name]

        # 将新工作表名称写入C1单元格中
        worksheet['C1'] = worksheet_name

        # 保存Excel文件
        workbook.save(file_path)

        # 更新工作表名称列表
        worksheet_names = workbook.sheetnames
        worksheet_name_combobox.configure(values=worksheet_names)

def on_worksheet_selected(event):
    selected_worksheet_name = worksheet_name_var.get()
    save_last_worksheet_name(selected_worksheet_name)

worksheet_name_combobox.bind('<<ComboboxSelected>>', on_worksheet_selected)


new_button = ttk.Button(root, style='my.TButton', text="新しいシートを制作", command=add_worksheet)
new_button.pack(pady=10)

# 1. 定义一个函数用于重新选择文件并加载
def reload_file():
    global file_path, workbook, worksheet_names
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    if file_path:
        save_file_path(file_path)
        workbook = openpyxl.load_workbook(file_path)
        worksheet_names = workbook.sheetnames
        worksheet_name_combobox.configure(values=worksheet_names)
        worksheet_name_var.set('シートを選択してください。')
    else:
        return

# 2. 创建一个新的按钮并将其添加到界面上
reload_button = ttk.Button(root, style='my.TButton', text="新しいExcelファイルを選択", command=reload_file)
reload_button.pack(pady=10)



def clear_amount_entry(event):
    amount.set('')

date = tk.StringVar(value='')
expense_type_var = tk.StringVar(value='昼食（未）')
amount = tk.StringVar(value='金額を入力してください。')


# 添加 back_to_selection 函数
def back_to_selection():
    global input_frame
    input_frame.destroy()

    # 重新创建并显示 worksheet_name_combobox
    global worksheet_name_combobox
    worksheet_name_combobox = Combobox(root, style='my.TCombobox', textvariable=worksheet_name_var, values=worksheet_names, width=30)
    worksheet_name_combobox.pack(pady=10)

    # 重新创建并显示 new_button
    global new_button
    new_button = ttk.Button(root, style='my.TButton', text="新しいシートを制作", command=add_worksheet)
    new_button.pack(pady=10)

    global reload_button
    reload_button = ttk.Button(root, style='my.TButton', text="新しいExcelファイルを選択", command=reload_file)
    reload_button.pack(pady=10)

    # 重新创建并显示 ok_button
    global ok_button
    ok_button = ttk.Button(root, style='my.TButton', text="次へ", command=ok_callback)
    ok_button.pack(pady=10)


def ok_callback():
    global worksheet_names, input_frame
    worksheet_name = worksheet_name_var.get()
    if worksheet_name not in worksheet_names:
        messagebox.showerror("エラーメッセージ", "無効のシート名")
        return

    else:
        worksheet = workbook[worksheet_name]
        worksheet_name_combobox.destroy()
        new_button.destroy()
        ok_button.destroy()
        reload_button.destroy()

        # 添加元素以輸入數據
        input_frame = ttk.Frame(root, style='my.TFrame', width=500, height=400)
        input_frame.pack_propagate(0)  # 禁止自动调整 Frame 的大小
        input_frame.pack()

        cal_frame = ttk.Frame(input_frame, width=30)
        cal_frame.pack(pady=10)
        cal = DateEntry(cal_frame, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd', textvariable=date)
        cal.pack(side='left')


        expense_type_combobox = Combobox(input_frame, style='my.TCombobox', width=30, textvariable=expense_type_var, values=['昼食（未）', '昼食（済）', '交通費（未）', '交通費（済）'])
        expense_type_combobox.pack(pady=10)

        def is_valid_input(input_string):
            allowed_chars = set('0123456789+-')
            return all(char in allowed_chars for char in input_string)

        def calculate():
            expression = amount.get()
            try:
                result = eval(expression)
                amount.set(str(result))
            except:
                amount.set('数字と数字の間に演算子を付けてください。')

        amount_entry = ttk.Entry(input_frame, style='my.TEntry', width=50, textvariable=amount)
        amount_entry.pack(pady=10)
        amount_entry.bind('<FocusIn>', clear_amount_entry)
        amount_entry.configure(validate="key", validatecommand=(amount_entry.register(is_valid_input), '%S'))

        calculate_button = ttk.Button(input_frame, style='my.TButton', text="計算", command=calculate)
        calculate_button.pack(pady=10)

        # 创建确认框
        confirm_var = tk.IntVar()
        confirm_checkbutton = ttk.Checkbutton(input_frame, style='my.TCheckbutton', text='複数の領収書がありますか？', variable=confirm_var)
        confirm_checkbutton.pack(pady=10)


    def submit():
        global date, expense_type_var, amount
        if not date.get():
            messagebox.showerror("エラーメッセージ", "日付を選択してください！")
            return
        if not expense_type_var.get():
            messagebox.showerror("エラーメッセージ", "金額種類を選択してください！")
            return
        if not amount.get() or amount.get() == '金額を入力してください。':
            messagebox.showerror("エラーメッセージ", "金額を入力してください！")
            return
        elif not amount.get().isdigit():
            messagebox.showerror("エラーメッセージ", "計算ボタンを押してください！")
            return

        data = [(date.get(), expense_type_var.get(), amount.get())]

        if data:
            for date, expense_type, amount in data:
                
               
            # 转换日期格式
                date_object = datetime.strptime(date, '%Y-%m-%d')
                formatted_date = date_object.strftime('%m/%d')

                # 写入数据到Excel文件
                row = 6
                while worksheet.cell(row=row, column=2).value is not None:
                    row += 1

                if row > 20:
                    row = 6
                    while worksheet.cell(row=row, column=6).value is not None:
                        row += 1
                    worksheet.cell(row=row, column=6).value = formatted_date
                    worksheet.cell(row=row, column=7).value = expense_type
                    worksheet.cell(row=row, column=8).value = float(amount)
                    worksheet.cell(row=row, column=8).number_format = '¥#,##0'
                    column = 6
                else:
                    while worksheet.cell(row=row, column=2).value is not None:
                        row += 1
                    worksheet.cell(row=row, column=2).value = formatted_date
                    worksheet.cell(row=row, column=3).value = expense_type
                    worksheet.cell(row=row, column=4).value = float(amount)
                    worksheet.cell(row=row, column=4).number_format = '¥#,##0'
                    column = 2
                if column == 2 or column == 6:
                    if confirm_var.get() == 1:
                        worksheet.cell(row=row, column=column).font = openpyxl.styles.Font(color="ff0000", size=14)
                    else:
                        worksheet.cell(row=row, column=column).font = openpyxl.styles.Font(color="000000", size=14)


            try:
                workbook.save(file_path)
                messagebox.showinfo("メッセージ", "データを入力完了です！")
            except PermissionError:
                messagebox.showerror("エラーメッセージ", "Excelを閉じってください!")
        input_frame.destroy()
        root.destroy()
        

    submit_button = ttk.Button(input_frame, style='my.TButton', text="入力完了", command=submit)
    submit_button.pack(pady=10)

    # 添加返回按钮
    back_button = ttk.Button(input_frame, style='my.TButton', text="戻る", command=back_to_selection)
    back_button.pack(pady=10)


ok_button = ttk.Button(root, style='my.TButton', text="次へ", command=ok_callback)
ok_button.pack(pady=10)


root.mainloop()
